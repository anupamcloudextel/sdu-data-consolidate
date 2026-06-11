from __future__ import annotations

from datetime import datetime
from pathlib import Path
import argparse
import io
import os
import re
import sys

import pandas as pd
from openpyxl import load_workbook

TPT_DASH_PATTERN = re.compile(r"-tpt", re.IGNORECASE)
TPT_SUFFIX_PATTERN = re.compile(r"tpt$", re.IGNORECASE)
NUMERIC_SUFFIX_PATTERN = re.compile(r"-\d+$")


# Map full circle names found in consolidated_hp.xlsx to ERPNext circle codes
# used in the existing "RSU Master.csv".
CIRCLE_TO_CODE: dict[str, str] = {
    "Chandigarh": "PUN",
    "Punjab": "PUN",
    "Delhi": "DEL",
    "Karnataka": "KTK",
    "Bangalore": "KTK",
    "Kolkata": "KOL",
    "Mumbai": "M&G",
    "M&G": "M&G",
    "Pune": "M&G",
    "Tamil Nadu": "TAM",
    "UPW": "UPW",
}

FIBER_ITEM_PREFIX = "SER-FSDU-CFT-"
FIBER_ORDER = ["6F", "12F", "24F", "48F"]

# ERPNext v15 TPT Status select options.
VALID_TPT_STATUS_VALUES: dict[str, str] = {
    "tpt route": "TPT Route",
    "cutover": "CutOver",
    "resign writing": "Resign Writing",
    "reparenting": "Reparenting",
    "ring connectivity": "Ring Connectivity",
}

OUTPUT_COLUMNS = [
    "ID",
    "RSU Code",
    "SDU Billing Rates",
    "TPT Status",
    "Circle",
    "Decom Date",
    "ID (Rsu Items)",
    "Item (Rsu Items)",
    "Plan FAT (Rsu Items)",
    "Plan Fibre length (Rsu Items)",
    "Plan HP (Rsu Items)",
]


def normalize_fiber_capacity(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip().upper().replace(" ", "")
    if not text:
        return ""
    if not text.endswith("F"):
        text = text + "F"
    return text


def map_circle_to_code(circle_text: str) -> str:
    text = circle_text.strip()
    return CIRCLE_TO_CODE.get(text, text)


def normalize_link_type_to_tpt_status(value: object) -> str:
    """Map consolidated ``Link Type`` values to ERPNext v15 TPT Status select values."""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if not text:
        return ""
    return VALID_TPT_STATUS_VALUES.get(re.sub(r"\s+", " ", text).strip().lower(), "")


def remove_tpt_suffix(value: str) -> str:
    without_dash_tpt = TPT_DASH_PATTERN.sub("", value)
    return TPT_SUFFIX_PATTERN.sub("", without_dash_tpt)


def normalize_cluster_id_for_match(value: object) -> str | None:
    """Normalize Cluster ID for lookup: strip whitespace, lowercase, drop TPT markers."""
    if pd.isna(value):
        return None
    normalized = re.sub(r"\s+", "", str(value).strip()).lower()
    if not normalized or normalized == "nan":
        return None
    normalized = remove_tpt_suffix(normalized)
    return normalized or None


def suffix_zero_key(cluster_id: str) -> str | None:
    if NUMERIC_SUFFIX_PATTERN.search(cluster_id):
        return None
    return f"{cluster_id}-0"


def ensure_rsu_code_text(value: object) -> str:
    """Keep RSU Code as plain text (avoids Excel scientific notation like 3.20E+01)."""
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except (TypeError, ValueError):
        pass
    text = str(value).strip()
    return text if text and text.lower() != "nan" else ""


def format_rsu_code_for_csv_export(value: object) -> str:
    """Format RSU Code for CSV so Excel keeps values like 04O-0 and 32E-0 as text."""
    text = ensure_rsu_code_text(value)
    if not text:
        return ""
    # Leading tab forces Excel to treat the cell as text when opening CSV.
    return f"\t{text}"


def resolve_cluster_to_rsu(cluster_id: str, mapping: dict[str, str]) -> str | None:
    normalized = normalize_cluster_id_for_match(cluster_id)
    if not normalized:
        return None
    if normalized in mapping:
        return mapping[normalized]
    zero_key = suffix_zero_key(normalized)
    if zero_key and zero_key in mapping:
        return mapping[zero_key]
    return None


def to_nullable_int(value: object) -> object:
    if value is None or value == "":
        return pd.NA
    try:
        if pd.isna(value):
            return pd.NA
    except (TypeError, ValueError):
        pass
    try:
        return int(round(float(value)))
    except (TypeError, ValueError):
        return pd.NA


def _excel_serial_to_datetime(value: object) -> object:
    """HOTO Date cells sometimes arrive as Excel serial numbers instead of dates."""
    if isinstance(value, (int, float)) and not pd.isna(value):
        try:
            return pd.Timestamp("1899-12-30") + pd.to_timedelta(float(value), unit="D")
        except (ValueError, OverflowError):
            return pd.NaT
    return value


def _read_bytes_with_shared_access_windows(path: Path) -> bytes:
    """Read a file's bytes on Windows using CreateFileW with full sharing
    flags so we can read even when Excel holds an exclusive write lock."""
    import ctypes
    from ctypes import wintypes

    GENERIC_READ = 0x80000000
    FILE_SHARE_READ = 0x00000001
    FILE_SHARE_WRITE = 0x00000002
    FILE_SHARE_DELETE = 0x00000004
    OPEN_EXISTING = 3
    INVALID_HANDLE_VALUE = ctypes.c_void_p(-1).value
    FILE_ATTRIBUTE_NORMAL = 0x80

    kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)

    kernel32.CreateFileW.restype = wintypes.HANDLE
    kernel32.CreateFileW.argtypes = [
        wintypes.LPCWSTR, wintypes.DWORD, wintypes.DWORD,
        ctypes.c_void_p, wintypes.DWORD, wintypes.DWORD, wintypes.HANDLE,
    ]
    kernel32.GetFileSizeEx.restype = wintypes.BOOL
    kernel32.GetFileSizeEx.argtypes = [wintypes.HANDLE, ctypes.POINTER(ctypes.c_longlong)]
    kernel32.ReadFile.restype = wintypes.BOOL
    kernel32.ReadFile.argtypes = [
        wintypes.HANDLE, ctypes.c_void_p, wintypes.DWORD,
        ctypes.POINTER(wintypes.DWORD), ctypes.c_void_p,
    ]
    kernel32.CloseHandle.restype = wintypes.BOOL
    kernel32.CloseHandle.argtypes = [wintypes.HANDLE]

    handle = kernel32.CreateFileW(
        str(path),
        GENERIC_READ,
        FILE_SHARE_READ | FILE_SHARE_WRITE | FILE_SHARE_DELETE,
        None,
        OPEN_EXISTING,
        FILE_ATTRIBUTE_NORMAL,
        None,
    )
    if not handle or handle == INVALID_HANDLE_VALUE:
        raise OSError(ctypes.get_last_error(), f"CreateFileW failed for {path}")

    try:
        file_size = ctypes.c_longlong(0)
        if not kernel32.GetFileSizeEx(handle, ctypes.byref(file_size)):
            raise OSError(ctypes.get_last_error(), "GetFileSizeEx failed")

        remaining = file_size.value
        chunks: list[bytes] = []
        chunk_size = 1024 * 1024
        buffer = ctypes.create_string_buffer(chunk_size)
        bytes_read = wintypes.DWORD(0)

        while remaining > 0:
            to_read = min(chunk_size, remaining)
            ok = kernel32.ReadFile(
                handle, buffer, to_read, ctypes.byref(bytes_read), None
            )
            if not ok:
                raise OSError(ctypes.get_last_error(), "ReadFile failed")
            if bytes_read.value == 0:
                break
            chunks.append(buffer.raw[: bytes_read.value])
            remaining -= bytes_read.value

        return b"".join(chunks)
    finally:
        kernel32.CloseHandle(handle)


def read_excel_lock_safe(
    input_path: Path,
    sheet_name: str | object = 0,
    dtype: object | None = None,
) -> pd.DataFrame:
    """Read an Excel sheet. Falls back to a Windows shared-read on
    PermissionError so the script works while the file is open in Excel."""
    read_kwargs: dict[str, object] = {"sheet_name": sheet_name}
    if dtype is not None:
        read_kwargs["dtype"] = dtype
    try:
        return pd.read_excel(input_path, **read_kwargs)
    except PermissionError:
        if os.name != "nt":
            raise
        print(
            f"'{input_path.name}' appears to be open (locked). "
            "Reading via a Windows shared-read."
        )
        data = _read_bytes_with_shared_access_windows(input_path)
        return pd.read_excel(io.BytesIO(data), **read_kwargs)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Convert consolidated_hp.xlsx into an ERPNext v15-importable RSU Master CSV "
            "(one parent row per RSU Code + TPT Status + 4 fiber child rows)."
        )
    )
    parser.add_argument("--input-file", default="consolidated_hp.xlsx")
    parser.add_argument("--input-sheet", default="Consolidated")
    parser.add_argument("--output-file", default="RSU Master Generated.csv")
    parser.add_argument(
        "--cluster-site-file",
        default="cluster_site_comparison.xlsx",
        help=(
            'Excel file with columns "Cluster ID" and "total match". '
            "RSU Code is taken from total match for Cluster IDs present in "
            "the consolidated input."
        ),
    )
    parser.add_argument(
        "--cluster-site-sheet",
        default=None,
        help="Sheet name in the cluster/site comparison file (default: first sheet).",
    )
    return parser.parse_args()


def load_cluster_to_rsu_mapping(
    comparison_path: Path, sheet_name: str | None
) -> dict[str, str]:
    """Read cluster_site_comparison.xlsx and return {normalized Cluster ID -> RSU Code}."""
    if not comparison_path.exists():
        print(f"Cluster/site comparison file not found: {comparison_path}")
        return {}

    effective_sheet: object = sheet_name if sheet_name is not None else 0
    df = read_excel_lock_safe(comparison_path, effective_sheet, dtype=str)

    cluster_col = next(
        (c for c in df.columns if str(c).strip().lower() == "cluster id"), None
    )
    rsu_col = next(
        (c for c in df.columns if str(c).strip().lower() == "total match"), None
    )
    if cluster_col is None or rsu_col is None:
        print(
            f"Comparison file '{comparison_path.name}' is missing required columns "
            "'Cluster ID' and/or 'total match'."
        )
        return {}

    mapping: dict[str, str] = {}
    for cluster_value, rsu_value in zip(df[cluster_col], df[rsu_col]):
        if pd.isna(cluster_value) or pd.isna(rsu_value):
            continue
        rsu_code = ensure_rsu_code_text(rsu_value)
        normalized_cluster = normalize_cluster_id_for_match(cluster_value)
        if normalized_cluster and rsu_code:
            mapping[normalized_cluster] = rsu_code
    return mapping


def apply_rsu_code_text_format(workbook_path: Path, column_name: str = "RSU Code") -> None:
    """Set Excel text format (@) on the RSU Code column so values like 32E-0 stay text."""
    workbook = load_workbook(workbook_path)
    worksheet = workbook.active
    header_cells = next(worksheet.iter_rows(min_row=1, max_row=1))
    rsu_column_index = next(
        (
            cell.column
            for cell in header_cells
            if str(cell.value).strip() == column_name
        ),
        None,
    )
    if rsu_column_index is None:
        workbook.close()
        return

    for row_index in range(2, worksheet.max_row + 1):
        cell = worksheet.cell(row=row_index, column=rsu_column_index)
        if cell.value in (None, ""):
            continue
        cell.value = ensure_rsu_code_text(cell.value)
        cell.number_format = "@"

    workbook.save(workbook_path)
    workbook.close()


def write_rsu_master_output(out_df: pd.DataFrame, output_path: Path) -> Path:
    """Write RSU Master output matching the RSU Master.csv layout for ERPNext import."""
    prepared = out_df.copy()
    prepared["RSU Code"] = prepared["RSU Code"].map(ensure_rsu_code_text)

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.suffix.lower() == ".xlsx":
        prepared.to_excel(output_path, index=False, na_rep="")
        apply_rsu_code_text_format(output_path)
        return output_path

    csv_ready = prepared.copy()
    csv_ready["RSU Code"] = csv_ready["RSU Code"].map(format_rsu_code_for_csv_export)
    csv_ready.to_csv(output_path, index=False, na_rep="")

    excel_copy_path = output_path.with_suffix(".xlsx")
    prepared.to_excel(excel_copy_path, index=False, na_rep="")
    apply_rsu_code_text_format(excel_copy_path)
    print(f"Excel copy (RSU Code as text): {excel_copy_path}")
    return output_path


def main() -> int:
    args = parse_args()
    input_path = Path(args.input_file).resolve()
    output_path = Path(args.output_file).resolve()

    if not input_path.exists():
        print(f"Input file not found: {input_path}")
        return 1

    df = read_excel_lock_safe(input_path, args.input_sheet)

    comparison_path = Path(args.cluster_site_file).resolve()
    cluster_to_rsu = load_cluster_to_rsu_mapping(comparison_path, args.cluster_site_sheet)
    if not cluster_to_rsu:
        print("No Cluster ID -> RSU Code mapping loaded. Cannot build RSU Master.")
        return 1

    print(
        f"Loaded {len(cluster_to_rsu)} Cluster ID -> RSU Code entries "
        f"from {comparison_path.name}."
    )

    required = [
        "Cluster ID",
        "Fiber Capacity",
        "Planned FAT",
        "Planned HP",
        "Fiber Length as per HOTO",
        "Circle",
    ]
    missing = [column for column in required if column not in df.columns]
    if missing:
        print(f"Missing required columns in input: {missing}")
        return 1

    df = df.copy()
    df = df.dropna(subset=["Cluster ID"])
    df["Cluster ID"] = df["Cluster ID"].astype(str).str.strip()
    df = df[(df["Cluster ID"] != "") & (df["Cluster ID"].str.lower() != "nan")]

    df["RSU Code"] = df["Cluster ID"].map(
        lambda cluster_id: resolve_cluster_to_rsu(cluster_id, cluster_to_rsu)
    )
    unmatched_mask = df["RSU Code"].isna()
    unmatched_clusters = sorted(df.loc[unmatched_mask, "Cluster ID"].unique().tolist())
    df = df[~unmatched_mask].copy()
    df["RSU Code"] = df["RSU Code"].map(ensure_rsu_code_text)
    df = df[df["RSU Code"] != ""].copy()
    if unmatched_clusters:
        print(
            f"Skipped {len(unmatched_clusters)} Cluster ID(s) with no RSU Code in "
            f"{comparison_path.name} (after normalized comparison)."
        )
        for cluster_id in unmatched_clusters:
            print(f"  {cluster_id!r}")

    if df.empty:
        print("No rows left after matching Cluster IDs to RSU Codes.")
        return 1

    if "Link Type" in df.columns:
        df["TPT Status"] = df["Link Type"].map(normalize_link_type_to_tpt_status)
    else:
        df["TPT Status"] = ""

    df["Fiber Capacity"] = df["Fiber Capacity"].map(normalize_fiber_capacity)
    df = df[df["Fiber Capacity"].isin(FIBER_ORDER)]

    df["Fiber Length as per HOTO"] = pd.to_numeric(
        df["Fiber Length as per HOTO"], errors="coerce"
    ).fillna(0)
    df["Planned FAT"] = pd.to_numeric(df["Planned FAT"], errors="coerce")
    df["Planned HP"] = pd.to_numeric(df["Planned HP"], errors="coerce")

    # Each unique (RSU Code, TPT Status) becomes its own 4-row block.
    df = df.sort_values(
        by=[
            "RSU Code",
            "TPT Status",
            "Fiber Capacity",
            "Fiber Length as per HOTO",
        ],
        ascending=[True, True, True, False],
    )
    df = df.drop_duplicates(
        subset=["RSU Code", "TPT Status", "Fiber Capacity"],
        keep="first",
    )

    rows: list[dict[str, object]] = []

    for (rsu_code, tpt_status_value), group in df.groupby(
        ["RSU Code", "TPT Status"], sort=True
    ):
        circle_values = group["Circle"].dropna().astype(str).str.strip()
        circle_values = circle_values[circle_values != ""]
        circle_text = circle_values.mode().iloc[0] if not circle_values.empty else ""
        circle_code = map_circle_to_code(circle_text)
        billing_rate = f"{circle_code} 1 Year Fix" if circle_code else ""

        group_by_capacity = group.set_index("Fiber Capacity")

        for index, capacity in enumerate(FIBER_ORDER):
            item_code = f"{FIBER_ITEM_PREFIX}{capacity}"

            if capacity in group_by_capacity.index:
                capacity_row = group_by_capacity.loc[capacity]
                if isinstance(capacity_row, pd.DataFrame):
                    capacity_row = capacity_row.iloc[0]
                plan_fat_value = capacity_row["Planned FAT"]
                plan_hp_value = capacity_row["Planned HP"]
            else:
                plan_fat_value = 0
                plan_hp_value = 0

            row: dict[str, object] = {column: "" for column in OUTPUT_COLUMNS}
            if index == 0:
                row["RSU Code"] = ensure_rsu_code_text(rsu_code)
                row["SDU Billing Rates"] = billing_rate
                row["TPT Status"] = tpt_status_value
                row["Circle"] = circle_code
            row["Item (Rsu Items)"] = item_code
            row["Plan FAT (Rsu Items)"] = plan_fat_value
            row["Plan Fibre length (Rsu Items)"] = 0
            row["Plan HP (Rsu Items)"] = plan_hp_value
            rows.append(row)

    out_df = pd.DataFrame(rows, columns=OUTPUT_COLUMNS)
    out_df["RSU Code"] = out_df["RSU Code"].astype(object)

    numeric_columns = [
        "Plan FAT (Rsu Items)",
        "Plan Fibre length (Rsu Items)",
        "Plan HP (Rsu Items)",
    ]
    for column in numeric_columns:
        out_df[column] = out_df[column].map(to_nullable_int).astype("Int64")

    try:
        output_path = write_rsu_master_output(out_df, output_path)
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_path = output_path.with_name(
            f"{output_path.stem}_{timestamp}{output_path.suffix}"
        )
        print(
            f"'{output_path.name}' is open or write-locked. "
            f"Writing to '{fallback_path.name}' instead."
        )
        output_path = write_rsu_master_output(out_df, fallback_path)

    group_count = len(out_df) // len(FIBER_ORDER)
    distinct_rsus = int(
        out_df["RSU Code"].astype(str).str.strip().replace("", pd.NA).dropna().nunique()
    )
    print(f"Output: {output_path}")
    print(f"(RSU Code x TPT Status) groups written: {group_count}")
    print(f"Distinct RSU Codes: {distinct_rsus}")
    print(f"Total rows (groups x {len(FIBER_ORDER)}): {len(out_df)}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
